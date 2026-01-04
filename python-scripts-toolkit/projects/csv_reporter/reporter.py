"""
CSV Report Generator - Reads CSV files and generates summary reports.

Supports advanced statistics including median, standard deviation, variance,
and percentiles (P25, P50, P75) using Python's statistics module.

Usage:
    python -m projects.csv_reporter.reporter input.csv
    python -m projects.csv_reporter.reporter input.csv --output report.txt
    python -m projects.csv_reporter.reporter input.csv --filter-column category --filter-value "Food"
    python -m projects.csv_reporter.reporter input.csv --date-from 2024-01-01 --date-to 2024-12-31
    python -m projects.csv_reporter.reporter input.csv --full-stats
    python -m projects.csv_reporter.reporter input.csv --stats median,stdev,p75
"""
import argparse
import csv
import statistics
from pathlib import Path
from collections import defaultdict
from glob import glob
from datetime import datetime
from typing import Optional, List, Dict, Any
from enum import Enum
from dataclasses import dataclass, field
import json
import sys

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from config import CSV_REPORTER_CONFIG, DATA_DIR
from utils.logger import setup_logger
from utils.helpers import parse_date



EXCEL_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb'}
CSV_EXTENSIONS = {'.csv', '.tsv', '.txt'}

# Check for openpyxl availability
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Check for matplotlib availability
try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend for headless servers
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    plt = None


class OutputFormat(Enum):
    """Supported output formats for reports."""
    TEXT = "text"
    JSON = "json"
    MARKDOWN = "markdown"
    HTML = "html"


@dataclass
class ReportMetadata:
    """Structured metadata for report generation.

    Contains source file information, generation timestamp, and row counts
    used across all output formats for consistent metadata representation.
    """
    sources: List[str] = field(default_factory=list)
    generated_at: str = ""
    total_rows: int = 0
    columns: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        """Convert metadata to dictionary for JSON serialization."""
        return {
            "sources": self.sources,
            "generated_at": self.generated_at,
            "total_rows": self.total_rows,
            "columns": self.columns,
        }


class ReportEncoder(json.JSONEncoder):
    """Custom JSON encoder for report data types.

    Handles serialization of datetime objects, Path objects, and other
    special types that may appear in report data.
    """

    def default(self, obj: Any) -> Any:
        """Convert special types to JSON-serializable formats."""
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, Path):
            return str(obj)
        if hasattr(obj, 'to_dict'):
            return obj.to_dict()
        return super().default(obj)


def _get_file_type(path: Path) -> str:
    """Detect file type by extension.

    Returns:
        'excel' for Excel files (.xlsx, .xls, .xlsm, .xlsb)
        'csv' for CSV files (.csv, .tsv, .txt)
        'unknown' for unrecognized extensions
    """
    suffix = path.suffix.lower()
    if suffix in EXCEL_EXTENSIONS:
        return 'excel'
    elif suffix in CSV_EXTENSIONS:
        return 'csv'
    return 'unknown'


class CSVReporter:
    """Generates reports from CSV/Excel data with aggregation and filtering."""

    # Available advanced statistics
    AVAILABLE_STATS = {
        "median": "Median value",
        "stdev": "Standard deviation",
        "variance": "Variance",
        "p25": "25th percentile (Q1)",
        "p50": "50th percentile (Q2/Median)",
        "p75": "75th percentile (Q3)",
    }

    # Available chart types
    CHART_TYPES = {
        "bar": "Vertical bar chart",
        "hbar": "Horizontal bar chart",
        "pie": "Pie chart for category breakdowns",
        "line": "Line chart for time-series data",
    }

    # Supported chart output formats
    CHART_FORMATS = {'.png', '.pdf', '.svg', '.jpg', '.jpeg'}

    # Default chart styling
    CHART_DEFAULTS = {
        "figsize": (10, 6),
        "dpi": 100,
        "bar_color": "#3498db",
        "pie_cmap": "Set3",
        "line_color": "#2ecc71",
        "title_fontsize": 14,
        "label_fontsize": 10,
    }

    def __init__(self, input_patterns: List[str]):
        self.input_paths = self._resolve_paths(input_patterns)
        self.logger = setup_logger("csv_reporter")
        self.data: List[Dict[str, Any]] = []
        self.headers: List[str] = []
        self.all_headers: List[str] = [] # FIX: Added to store all unique headers across merged files
        self.numeric_columns: List[str] = []
        self.date_column: Optional[str] = None
        self.category_column: Optional[str] = None
        # Statistics configuration
        self.full_stats: bool = False
        self.selected_stats: Optional[List[str]] = None

    def configure_stats(self, full_stats: bool = False, stats_list: Optional[str] = None) -> None:
        """Configure which advanced statistics to display in reports.

        Args:
            full_stats: If True, display all available advanced statistics
            stats_list: Comma-separated list of specific stats to show
                        (e.g., "median,stdev,p75")
        """
        self.full_stats = full_stats
        if stats_list:
            requested = [s.strip().lower() for s in stats_list.split(",")]
            valid_stats = []
            for stat in requested:
                if stat in self.AVAILABLE_STATS:
                    valid_stats.append(stat)
                else:
                    self.logger.warning("Unknown statistic: %s (available: %s)",
                                       stat, ", ".join(self.AVAILABLE_STATS.keys()))
            self.selected_stats = valid_stats if valid_stats else None
        else:
            self.selected_stats = None

    def _resolve_paths(self, patterns: List[str]) -> List[Path]:
        """Resolve glob patterns to existing file paths."""
        seen = set()
        files = []
        for p in patterns:
            # First, try to resolve relative to the current working directory
            for f in glob(p):
                path = Path(f).resolve()
                if path not in seen and path.is_file():
                    seen.add(path)
                    files.append(Path(f))
            # Then, try to resolve relative to DATA_DIR
            for f in glob(str(DATA_DIR / p)):
                path = Path(f).resolve()
                if path not in seen and path.is_file():
                    seen.add(path)
                    files.append(Path(f))
        return files

    def _load_excel(self, path: Path, sheet_name: Optional[str] = None) -> tuple:
        """Load data from an Excel file.

        Args:
            path: Path to the Excel file
            sheet_name: Name of the sheet to load (default: first sheet)

        Returns:
            Tuple of (headers, data) where headers is a list of column names
            and data is a list of dicts representing rows.

        Raises:
            ImportError: If openpyxl is not installed
            ValueError: If the specified sheet doesn't exist
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                f"Excel file support requires openpyxl. Install with:\n"
                f"  pip install openpyxl\n"
                f"Or convert '{path.name}' to CSV format."
            )

        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)

        # Select the sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read headers from first row
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return [], []

        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(header_row)]

        # Read data rows
        data = []
        for row in rows:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert None to empty string for consistency with CSV behavior
                    row_dict[headers[i]] = str(value) if value is not None else ""
            data.append(row_dict)

        wb.close()
        return headers, data

    def get_sheet_names(self, path: Path) -> List[str]:
        """Get list of sheet names from an Excel file.

        Args:
            path: Path to the Excel file

        Returns:
            List of sheet names

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                f"Excel file support requires openpyxl. Install with:\n"
                f"  pip install openpyxl"
            )

        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names

    def _load_single_file(self, path: Path, sheet_name: Optional[str] = None) -> tuple:
        """Load a single file (CSV or Excel) and return headers and data.

        Args:
            path: Path to the file
            sheet_name: Sheet name for Excel files (ignored for CSV)

        Returns:
            Tuple of (headers, data)
        """
        file_type = _get_file_type(path)

        if file_type == 'excel':
            return self._load_excel(path, sheet_name)
        else:
            # Default to CSV loading
            with open(path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                headers = list(reader.fieldnames or [])
                data = list(reader)
            return headers, data

    def load(self, merge_strategy: str = "append", join_key: Optional[str] = None, dedupe: bool = False, sheet_name: Optional[str] = None) -> bool:
        """Load and parse the CSV/Excel file(s) based on merge strategy.

        Args:
            merge_strategy: How to combine multiple files ('append' or 'join')
            join_key: Column name for join strategy
            dedupe: Remove duplicate rows
            sheet_name: Sheet name for Excel files (uses first sheet if not specified)

        Returns:
            True if loading succeeded, False otherwise
        """
        if not self.input_paths:
            self.logger.error("No input files found.")
            return False

        try:
            if len(self.input_paths) == 1:
                # Single file case
                self.headers, self.data = self._load_single_file(self.input_paths[0], sheet_name)
                self.all_headers = list(self.headers)
                self.logger.info("Loaded %d rows from %s", len(self.data), self.input_paths[0].name)
            else:
                # Multiple files case
                if merge_strategy == "append":
                    all_data = []
                    all_unique_headers = set()
                    for path in self.input_paths:
                        current_headers, rows = self._load_single_file(path, sheet_name)
                        all_unique_headers.update(current_headers)
                        all_data.extend(rows)
                        self.logger.info("Appended %d rows from %s", len(rows), path.name)
                    self.headers = sorted(list(all_unique_headers))
                    self.all_headers = list(self.headers)
                    self.data = all_data
                    self.logger.info("Total loaded %d rows from %d files using 'append' strategy.", len(self.data), len(self.input_paths))

                elif merge_strategy == "join":
                    if not join_key:
                        self.logger.error("Join strategy requires a --join-key.")
                        return False

                    # Check if join_key exists in all files
                    for p in self.input_paths:
                        file_headers, _ = self._load_single_file(p, sheet_name)
                        if join_key not in file_headers:
                            self.logger.error("Join key '%s' not found in file: %s", join_key, p.name)
                            return False

                    joined_data: Dict[str, Dict[str, Any]] = {}
                    all_unique_headers = set()

                    for i, path in enumerate(self.input_paths):
                        current_headers, rows = self._load_single_file(path, sheet_name)
                        all_unique_headers.update(current_headers)
                        for row in rows:
                            key_value = row.get(join_key)
                            if key_value:
                                if key_value not in joined_data:
                                    joined_data[key_value] = {join_key: key_value}
                                # Prefix columns from subsequent files to avoid conflicts
                                prefix = f"file{i+1}_" if i > 0 else ""
                                for header, value in row.items():
                                    if header != join_key:
                                        joined_data[key_value][f"{prefix}{header}"] = value
                                        all_unique_headers.add(f"{prefix}{header}")
                                    else:
                                        joined_data[key_value][header] = value
                    self.data = list(joined_data.values())
                    self.headers = sorted(list(all_unique_headers))
                    self.all_headers = list(self.headers)
                    self.logger.info("Total loaded %d rows from %d files using 'join' strategy on key '%s'.", len(self.data), len(self.input_paths), join_key)

            # Deduplicate rows if requested
            if dedupe and self.data:
                original_count = len(self.data)
                seen = set()
                unique_data = []
                for row in self.data:
                    # Create a hashable key from sorted row items
                    row_key = tuple(sorted((k, str(v)) for k, v in row.items()))
                    if row_key not in seen:
                        seen.add(row_key)
                        unique_data.append(row)
                self.data = unique_data
                removed = original_count - len(self.data)
                if removed > 0:
                    self.logger.info("Removed %d duplicate rows.", removed)

            self._detect_column_types()
            return True
        except Exception as e:
            self.logger.error("Error reading CSV: %s", e)
            return False

    def _detect_column_types(self) -> None:
        """Auto-detect column types for aggregation."""
        # Detect date column
        for col in CSV_REPORTER_CONFIG["date_columns"]:
            if col in self.headers:
                self.date_column = col
                break

        # Detect numeric columns
        for col in self.headers:
            if self._is_numeric_column(col):
                self.numeric_columns.append(col)

        # Detect category column
        for col in CSV_REPORTER_CONFIG["category_columns"]:
            if col in self.headers:
                self.category_column = col
                break

    def _is_numeric_column(self, column: str) -> bool:
        """Check if a column contains numeric data."""
        sample = [row.get(column, "") for row in self.data[:10]]
        try:
            for val in sample:
                if val.strip():
                    float(val.replace(",", "").replace("$", ""))
            return True
        except ValueError:
            return False

    def _parse_numeric(self, value: str) -> float:
        """Parse a numeric string, handling currency and commas."""
        if not value or not value.strip():
            return 0.0
        cleaned = value.replace(",", "").replace("$", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    def _compute_advanced_stats(self, values: List[float]) -> Dict[str, float]:
        """Compute advanced statistics for a list of numeric values.

        Args:
            values: List of float values to analyze

        Returns:
            Dictionary with calculated statistics (median, stdev, variance, percentiles)
        """
        result = {}

        if not values:
            return result

        # Median - works with any number of values
        result["median"] = statistics.median(values)

        # Standard deviation and variance require at least 2 values
        if len(values) >= 2:
            try:
                result["stdev"] = statistics.stdev(values)
                result["variance"] = statistics.variance(values)
            except statistics.StatisticsError:
                result["stdev"] = 0.0
                result["variance"] = 0.0
        else:
            result["stdev"] = 0.0
            result["variance"] = 0.0

        # Percentiles require at least 4 values for quartiles
        if len(values) >= 4:
            try:
                quantiles = statistics.quantiles(values, n=4)
                result["p25"] = quantiles[0]
                result["p50"] = quantiles[1]
                result["p75"] = quantiles[2]
            except statistics.StatisticsError:
                result["p25"] = result["median"]
                result["p50"] = result["median"]
                result["p75"] = result["median"]
        else:
            # For small datasets, use median for all percentiles
            result["p25"] = result["median"]
            result["p50"] = result["median"]
            result["p75"] = result["median"]

        return result

    def _prepare_chart_data(
        self,
        data: List[Dict[str, Any]],
        group_by: Optional[str] = None,
        value_column: Optional[str] = None
    ) -> tuple:
        """Prepare data for chart generation by aggregating values.

        Args:
            data: List of data rows to process
            group_by: Column to group data by (categories/labels)
            value_column: Numeric column to aggregate (values)

        Returns:
            Tuple of (labels, values, title) for chart rendering
        """
        # Determine the grouping column
        if group_by and group_by in self.headers:
            group_col = group_by
        elif self.category_column:
            group_col = self.category_column
        else:
            group_col = self.headers[0] if self.headers else None

        # Determine the value column
        if value_column and value_column in self.numeric_columns:
            val_col = value_column
        elif self.numeric_columns:
            val_col = self.numeric_columns[0]
        else:
            val_col = None

        if not group_col or not val_col:
            return [], [], "No data available for chart"

        # Aggregate data by group
        groups = defaultdict(float)
        for row in data:
            key = row.get(group_col, "Unknown")
            groups[key] += self._parse_numeric(row.get(val_col, ""))

        # Sort by value descending
        sorted_groups = sorted(groups.items(), key=lambda x: -x[1])

        labels = [item[0] for item in sorted_groups]
        values = [item[1] for item in sorted_groups]
        title = f"{val_col} by {group_col}"

        return labels, values, title

    def _create_bar_chart(
        self,
        labels: List[str],
        values: List[float],
        title: str,
        output_path: Path
    ) -> bool:
        """Create a vertical bar chart.

        Args:
            labels: Category labels for x-axis
            values: Numeric values for bar heights
            title: Chart title
            output_path: Path to save the chart

        Returns:
            True if chart was created successfully
        """
        if not HAS_MATPLOTLIB:
            return False

        fig, ax = plt.subplots(figsize=self.CHART_DEFAULTS["figsize"])

        bars = ax.bar(labels, values, color=self.CHART_DEFAULTS["bar_color"])

        ax.set_title(title, fontsize=self.CHART_DEFAULTS["title_fontsize"])
        ax.set_xlabel("Category", fontsize=self.CHART_DEFAULTS["label_fontsize"])
        ax.set_ylabel("Value", fontsize=self.CHART_DEFAULTS["label_fontsize"])

        # Rotate labels if there are many categories
        if len(labels) > 5:
            plt.xticks(rotation=45, ha='right')

        # Add value labels on bars
        for bar, val in zip(bars, values):
            height = bar.get_height()
            ax.annotate(f'{val:,.0f}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom',
                        fontsize=8)

        plt.tight_layout()
        plt.savefig(output_path, dpi=self.CHART_DEFAULTS["dpi"])
        plt.close(fig)

        return True

    def _create_horizontal_bar_chart(
        self,
        labels: List[str],
        values: List[float],
        title: str,
        output_path: Path
    ) -> bool:
        """Create a horizontal bar chart.

        Args:
            labels: Category labels for y-axis
            values: Numeric values for bar widths
            title: Chart title
            output_path: Path to save the chart

        Returns:
            True if chart was created successfully
        """
        if not HAS_MATPLOTLIB:
            return False

        fig, ax = plt.subplots(figsize=self.CHART_DEFAULTS["figsize"])

        # Reverse order so highest value is at top
        y_pos = range(len(labels))
        bars = ax.barh(y_pos, values[::-1], color=self.CHART_DEFAULTS["bar_color"])

        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels[::-1])
        ax.set_title(title, fontsize=self.CHART_DEFAULTS["title_fontsize"])
        ax.set_xlabel("Value", fontsize=self.CHART_DEFAULTS["label_fontsize"])

        # Add value labels on bars
        for bar, val in zip(bars, values[::-1]):
            width = bar.get_width()
            ax.annotate(f'{val:,.0f}',
                        xy=(width, bar.get_y() + bar.get_height() / 2),
                        xytext=(3, 0),
                        textcoords="offset points",
                        ha='left', va='center',
                        fontsize=8)

        plt.tight_layout()
        plt.savefig(output_path, dpi=self.CHART_DEFAULTS["dpi"])
        plt.close(fig)

        return True

    def _create_pie_chart(
        self,
        labels: List[str],
        values: List[float],
        title: str,
        output_path: Path
    ) -> bool:
        """Create a pie chart.

        Args:
            labels: Category labels for slices
            values: Numeric values for slice sizes
            title: Chart title
            output_path: Path to save the chart

        Returns:
            True if chart was created successfully
        """
        if not HAS_MATPLOTLIB:
            return False

        fig, ax = plt.subplots(figsize=self.CHART_DEFAULTS["figsize"])

        # Calculate percentages for display
        total = sum(values)
        percentages = [(v / total) * 100 if total > 0 else 0 for v in values]

        # Limit to top categories if too many (combine rest into "Other")
        max_slices = 8
        if len(labels) > max_slices:
            top_labels = labels[:max_slices - 1] + ["Other"]
            top_values = values[:max_slices - 1] + [sum(values[max_slices - 1:])]
        else:
            top_labels = labels
            top_values = values

        # Create pie chart with percentage labels
        wedges, texts, autotexts = ax.pie(
            top_values,
            labels=top_labels,
            autopct='%1.1f%%',
            colors=plt.cm.get_cmap(self.CHART_DEFAULTS["pie_cmap"]).colors,
            startangle=90
        )

        ax.set_title(title, fontsize=self.CHART_DEFAULTS["title_fontsize"])

        plt.tight_layout()
        plt.savefig(output_path, dpi=self.CHART_DEFAULTS["dpi"])
        plt.close(fig)

        return True

    def _create_line_chart(
        self,
        labels: List[str],
        values: List[float],
        title: str,
        output_path: Path
    ) -> bool:
        """Create a line chart for time-series or sequential data.

        Args:
            labels: Category/time labels for x-axis
            values: Numeric values for the line
            title: Chart title
            output_path: Path to save the chart

        Returns:
            True if chart was created successfully
        """
        if not HAS_MATPLOTLIB:
            return False

        fig, ax = plt.subplots(figsize=self.CHART_DEFAULTS["figsize"])

        # Plot line with markers
        ax.plot(
            labels,
            values,
            color=self.CHART_DEFAULTS["line_color"],
            marker='o',
            linewidth=2,
            markersize=6
        )

        ax.set_title(title, fontsize=self.CHART_DEFAULTS["title_fontsize"])
        ax.set_xlabel("Category", fontsize=self.CHART_DEFAULTS["label_fontsize"])
        ax.set_ylabel("Value", fontsize=self.CHART_DEFAULTS["label_fontsize"])

        # Rotate labels if there are many points
        if len(labels) > 5:
            plt.xticks(rotation=45, ha='right')

        # Add grid for readability
        ax.grid(True, linestyle='--', alpha=0.7)

        plt.tight_layout()
        plt.savefig(output_path, dpi=self.CHART_DEFAULTS["dpi"])
        plt.close(fig)

        return True

    def generate_chart(
        self,
        data: Optional[List[Dict[str, Any]]] = None,
        chart_type: str = "bar",
        output_path: Optional[Path] = None,
        group_by: Optional[str] = None,
        value_column: Optional[str] = None
    ) -> Optional[Path]:
        """Generate a chart visualization from the data.

        Args:
            data: Data to visualize (defaults to self.data)
            chart_type: Type of chart ('bar', 'hbar', 'pie', 'line')
            output_path: Where to save the chart (auto-generated if None)
            group_by: Column to group data by
            value_column: Numeric column to aggregate

        Returns:
            Path to generated chart file, or None if generation failed
        """
        if not HAS_MATPLOTLIB:
            self.logger.error(
                "Chart generation requires matplotlib. Install with:\n"
                "  pip install matplotlib"
            )
            return None

        # Validate chart type
        if chart_type not in self.CHART_TYPES:
            self.logger.error(
                "Unknown chart type: %s (available: %s)",
                chart_type, ", ".join(self.CHART_TYPES.keys())
            )
            return None

        # Use provided data or default
        chart_data = data if data is not None else self.data

        if not chart_data:
            self.logger.error("No data available for chart generation")
            return None

        # Prepare chart data
        labels, values, title = self._prepare_chart_data(chart_data, group_by, value_column)

        if not labels or not values:
            self.logger.error("Could not prepare data for chart")
            return None

        # Generate default output path if not provided
        if output_path is None:
            base_name = self.input_paths[0].stem if self.input_paths else "chart"
            output_path = Path(f"{base_name}_{chart_type}_chart.png")

        # Validate output format
        if output_path.suffix.lower() not in self.CHART_FORMATS:
            self.logger.warning(
                "Unknown format '%s', using PNG",
                output_path.suffix
            )
            output_path = output_path.with_suffix('.png')

        # Create the appropriate chart type
        chart_methods = {
            "bar": self._create_bar_chart,
            "hbar": self._create_horizontal_bar_chart,
            "pie": self._create_pie_chart,
            "line": self._create_line_chart,
        }

        success = chart_methods[chart_type](labels, values, title, output_path)

        if success:
            self.logger.info("Chart saved to: %s", output_path)
            return output_path
        else:
            self.logger.error("Failed to create chart")
            return None


    def _prepare_report_data(
        self,
        data: Optional[List[Dict[str, Any]]] = None,
        group_by: Optional[str] = None
    ) -> Dict[str, Any]:
        """Prepare structured report data for all output formats.

        Args:
            data: Data rows to include in report (uses self.data if None)
            group_by: Optional column name to group data by

        Returns:
            Dictionary with metadata, statistics, and group data ready for
            formatting into any output format (text, JSON, Markdown, HTML).
        """
        data = data or self.data

        # Build metadata
        metadata = ReportMetadata(
            sources=[p.name for p in self.input_paths],
            generated_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_rows=len(data),
            columns=self.headers
        )

        # Prepare the result structure
        result: Dict[str, Any] = {
            "metadata": metadata,
            "statistics": {},
            "groups": {},
            "category_breakdown": {}
        }

        # Compute statistics for numeric columns
        for col in self.numeric_columns:
            values = [self._parse_numeric(row.get(col, "")) for row in data]
            if values:
                col_stats = {
                    "total": sum(values),
                    "average": sum(values) / len(values) if values else 0,
                    "min": min(values),
                    "max": max(values),
                    "count": len([v for v in values if v != 0])
                }

                # Add advanced statistics if configured
                if self.full_stats or self.selected_stats:
                    advanced = self._compute_advanced_stats(values)
                    stats_to_show = self.selected_stats or list(self.AVAILABLE_STATS.keys())
                    for stat in stats_to_show:
                        if stat in advanced:
                            col_stats[stat] = advanced[stat]

                result["statistics"][col] = col_stats

        # Group by analysis
        if group_by and group_by in self.headers:
            groups = defaultdict(list)
            for row in data:
                key = row.get(group_by, "Unknown")
                groups[key].append(row)

            for group_name, group_data in sorted(groups.items()):
                group_stats = {"count": len(group_data)}
                for col in self.numeric_columns:
                    values = [self._parse_numeric(row.get(col, "")) for row in group_data]
                    if values:
                        group_stats[f"{col}_total"] = sum(values)
                result["groups"][group_name] = group_stats

        # Category breakdown (if detected and no group_by)
        elif self.category_column:
            categories = defaultdict(int)
            category_sums = defaultdict(lambda: defaultdict(float))

            for row in data:
                cat = row.get(self.category_column, "Unknown")
                categories[cat] += 1
                for col in self.numeric_columns:
                    category_sums[cat][col] += self._parse_numeric(row.get(col, ""))

            for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
                result["category_breakdown"][cat] = {
                    "count": count,
                    **{col: total for col, total in category_sums[cat].items()}
                }

        return result


    def _format_statistics_for_json(self, report_data: Dict[str, Any]) -> Dict[str, Any]:
        """Format report data for JSON output.

        Converts ReportMetadata to dict and ensures all values are
        JSON-serializable.

        Args:
            report_data: Dictionary from _prepare_report_data

        Returns:
            Dictionary with all values ready for JSON serialization.
        """
        result = {
            "metadata": report_data["metadata"].to_dict(),
            "statistics": {},
            "groups": report_data.get("groups", {}),
            "category_breakdown": report_data.get("category_breakdown", {})
        }

        # Format statistics with consistent precision
        for col, stats in report_data.get("statistics", {}).items():
            result["statistics"][col] = {
                key: round(val, 2) if isinstance(val, float) else val
                for key, val in stats.items()
            }

        return result

    def filter_data(
        self,
        filter_column: Optional[str] = None,
        filter_value: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Filter data based on criteria."""
        filtered = self.data.copy()

        # Filter by column value
        if filter_column and filter_value:
            filtered = [
                row for row in filtered
                if row.get(filter_column, "").lower() == filter_value.lower()
            ]

        # Filter by date range
        if self.date_column and (date_from or date_to):
            date_filtered = []
            from_date = parse_date(date_from) if date_from else None
            to_date = parse_date(date_to) if date_to else None

            for row in filtered:
                try:
                    row_date = parse_date(row.get(self.date_column, ""))
                    if from_date and row_date < from_date:
                        continue
                    if to_date and row_date > to_date:
                        continue
                    date_filtered.append(row)
                except ValueError:
                    continue

            filtered = date_filtered

        return filtered

    def generate_report(
        self,
        data: Optional[List[Dict[str, Any]]] = None,
        group_by: Optional[str] = None
    ) -> str:
        """Generate a summary report."""
        data = data or self.data

        lines = [
            "=" * 60, # FIX: Changed to show all input file names
            f"CSV REPORT: {', '.join([p.name for p in self.input_paths])}", # FIX: Show all input file names
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 60,
            "",
            f"Total Rows: {len(data)}",
            f"Columns: {', '.join(self.headers)}",
            "",
        ]

        # Numeric summaries
        if self.numeric_columns:
            lines.append("-" * 40)
            lines.append("NUMERIC SUMMARIES")
            lines.append("-" * 40)

            for col in self.numeric_columns:
                values = [self._parse_numeric(row.get(col, "")) for row in data]
                if values:
                    lines.append(f"\n{col}:")
                    lines.append(f"  Total:   {sum(values):,.2f}")
                    lines.append(f"  Average: {sum(values)/len(values):,.2f}")

                    # Add advanced statistics if configured
                    if self.full_stats or self.selected_stats:
                        advanced = self._compute_advanced_stats(values)
                        stats_to_show = self.selected_stats or list(self.AVAILABLE_STATS.keys())

                        if "median" in stats_to_show and "median" in advanced:
                            lines.append(f"  Median:  {advanced['median']:,.2f}")
                        if "stdev" in stats_to_show and "stdev" in advanced:
                            lines.append(f"  Std Dev: {advanced['stdev']:,.2f}")
                        if "variance" in stats_to_show and "variance" in advanced:
                            lines.append(f"  Variance:{advanced['variance']:,.2f}")

                    lines.append(f"  Min:     {min(values):,.2f}")
                    lines.append(f"  Max:     {max(values):,.2f}")

                    # Add percentiles after min/max if configured
                    if self.full_stats or self.selected_stats:
                        advanced = self._compute_advanced_stats(values)
                        stats_to_show = self.selected_stats or list(self.AVAILABLE_STATS.keys())

                        if "p25" in stats_to_show and "p25" in advanced:
                            lines.append(f"  P25:     {advanced['p25']:,.2f}")
                        if "p50" in stats_to_show and "p50" in advanced:
                            lines.append(f"  P50:     {advanced['p50']:,.2f}")
                        if "p75" in stats_to_show and "p75" in advanced:
                            lines.append(f"  P75:     {advanced['p75']:,.2f}")

                    lines.append(f"  Count:   {len([v for v in values if v != 0])}")

        # Group by analysis
        if group_by and group_by in self.headers:
            lines.append("")
            lines.append("-" * 40)
            lines.append(f"GROUPED BY: {group_by}")
            lines.append("-" * 40)

            groups = defaultdict(list)
            for row in data:
                key = row.get(group_by, "Unknown")
                groups[key].append(row)

            for group_name, group_data in sorted(groups.items()):
                lines.append(f"\n{group_name}: {len(group_data)} items")

                for col in self.numeric_columns:
                    values = [self._parse_numeric(row.get(col, "")) for row in group_data]
                    if values:
                        lines.append(f"  {col} total: {sum(values):,.2f}")

        # Category breakdown (if detected)
        elif self.category_column:
            lines.append("")
            lines.append("-" * 40)
            lines.append(f"BY {self.category_column.upper()}")
            lines.append("-" * 40)

            categories = defaultdict(int)
            category_sums = defaultdict(lambda: defaultdict(float))

            for row in data:
                cat = row.get(self.category_column, "Unknown")
                categories[cat] += 1
                for col in self.numeric_columns:
                    category_sums[cat][col] += self._parse_numeric(row.get(col, ""))

            for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
                lines.append(f"\n{cat}: {count} items")
                for col, total in category_sums[cat].items():
                    lines.append(f"  {col}: {total:,.2f}")

        lines.append("")
        lines.append("=" * 60)
        return "\n".join(lines)

    def export_summary_csv(self, output_path: Path, group_by: str, data: List[Dict[str, Any]]) -> bool:
        """Export a summary CSV grouped by a column."""
        if group_by not in self.headers:
            self.logger.error("Column not found: %s", group_by)
            return False
        
        groups = defaultdict(list)
        for row in data:
            key = row.get(group_by, "Unknown")
            groups[key].append(row)

        summary_data = []
        for group_name, group_data in groups.items():
            row = {group_by: group_name, "count": len(group_data)}
            for col in self.numeric_columns:
                values = [self._parse_numeric(r.get(col, "")) for r in group_data]
                row[f"{col}_total"] = sum(values)
                row[f"{col}_avg"] = sum(values) / len(values) if values else 0
            summary_data.append(row)

        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                if summary_data:
                    writer = csv.DictWriter(f, fieldnames=summary_data[0].keys())
                    writer.writeheader()
                    writer.writerows(summary_data)
            self.logger.info("Summary exported to: %s", output_path)
            return True
        except Exception as e:
            self.logger.error("Error writing CSV: %s", e)
            return False


def main():
    parser = argparse.ArgumentParser(
        description="Generate reports from CSV/Excel data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic single-file report
  reporter.py expenses.csv
  # Excel file (uses first sheet)
  reporter.py data.xlsx
  # Excel file with specific sheet
  reporter.py data.xlsx --sheet "Sales Data"
  # List available sheets in Excel file
  reporter.py data.xlsx --list-sheets
  # Multiple CSV files (append rows)
  reporter.py jan.csv feb.csv mar.csv
  # Using glob pattern
  reporter.py *.csv
  # Append merge explicitly
  reporter.py *.csv --merge append
  # Join CSVs on a common column
  reporter.py users.csv orders.csv --merge join --join-key user_id
  # Grouped report
  reporter.py *.csv --group-by category
  # Filter by column value
  reporter.py expenses.csv --filter-column type --filter-value Food
  # Date range filtering
  reporter.py expenses.csv --date-from 2024-01-01 --date-to 2024-06-30
  # Export grouped summary as CSV
  reporter.py *.csv --group-by category --export-csv summary.csv
  # Remove duplicate rows
  reporter.py *.csv --merge append --dedupe
  # Show all advanced statistics (median, std dev, variance, percentiles)
  reporter.py expenses.csv --full-stats
  # Show specific statistics only
  reporter.py expenses.csv --stats median,stdev,p75"""
    )
    parser.add_argument("inputs", nargs="+", help="CSV/Excel files or glob patterns")
    parser.add_argument("--output", "-o", type=Path, help="Output file for report")
    parser.add_argument("--group-by", "-g", help="Column to group by")
    parser.add_argument("--filter-column", "-fc", help="Column to filter on")
    parser.add_argument("--filter-value", "-fv", help="Value to filter for")
    parser.add_argument("--date-from", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--date-to", help="End date (YYYY-MM-DD)")
    parser.add_argument("--export-csv", type=Path, help="Export summary as CSV")
    parser.add_argument("--merge", choices=["append", "join"], default="append")
    parser.add_argument("--join-key", help="Required for join")
    parser.add_argument("--dedupe", action="store_true", help="Remove duplicate rows")
    parser.add_argument("--sheet", "-s", help="Excel sheet name (default: first sheet)")
    parser.add_argument("--list-sheets", action="store_true", help="List available sheets in Excel file")
    parser.add_argument("--full-stats", action="store_true",
                        help="Show all advanced statistics (median, std dev, variance, percentiles)")
    parser.add_argument("--stats", metavar="STATS",
                        help="Comma-separated list of stats to show (median,stdev,variance,p25,p50,p75)")
    parser.add_argument("--chart", action="store_true",
                        help="Generate a chart alongside the text report")
    parser.add_argument("--chart-type", choices=["bar", "hbar", "pie", "line"], default="bar",
                        help="Chart type: bar (default), hbar, pie, line")
    parser.add_argument("--chart-output", type=Path, metavar="FILE",
                        help="Output file for chart (PNG, PDF, SVG, JPG)")
    parser.add_argument("--chart-column", metavar="COLUMN",
                        help="Numeric column to visualize (default: first numeric column)")
    args = parser.parse_args()

    reporter = CSVReporter(args.inputs)

    # Handle --list-sheets
    if args.list_sheets:
        for path in reporter.input_paths:
            if _get_file_type(path) == 'excel':
                try:
                    sheets = reporter.get_sheet_names(path)
                    print(f"\n{path.name}:")
                    for i, sheet in enumerate(sheets, 1):
                        print(f"  {i}. {sheet}")
                except ImportError as e:
                    print(f"Error: {e}")
                    sys.exit(1)
            else:
                print(f"\n{path.name}: Not an Excel file")
        sys.exit(0)

    if not reporter.load(args.merge, args.join_key, args.dedupe, args.sheet):
        sys.exit(1)

    # Configure advanced statistics
    reporter.configure_stats(full_stats=args.full_stats, stats_list=args.stats)

    # Filter data
    filtered_data = reporter.filter_data(
        filter_column=args.filter_column,
        filter_value=args.filter_value,
        date_from=args.date_from,
        date_to=args.date_to
    )

    # Generate report
    report = reporter.generate_report(filtered_data, group_by=args.group_by)

    # Output
    if args.output:
        with open(args.output, 'w') as f:
            f.write(report)
        print(f"Report saved to: {args.output}")
    else:
        print(report)

    # Export summary CSV
    if args.export_csv and args.group_by:
        reporter.export_summary_csv(args.export_csv, args.group_by, filtered_data)

    # Generate chart
    if args.chart:
        chart_path = reporter.generate_chart(
            data=filtered_data,
            chart_type=args.chart_type,
            output_path=args.chart_output,
            group_by=args.group_by,
            value_column=args.chart_column
        )
        if chart_path:
            print(f"Chart saved to: {chart_path}")


if __name__ == "__main__":
    main()
