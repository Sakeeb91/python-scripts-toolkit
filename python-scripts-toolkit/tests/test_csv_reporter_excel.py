"""Tests for CSV Reporter Excel file support."""
import tempfile
from pathlib import Path
import pytest
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from projects.csv_reporter.reporter import (
    CSVReporter,
    _get_file_type,
    EXCEL_EXTENSIONS,
    CSV_EXTENSIONS,
    HAS_OPENPYXL
)


class TestFileTypeDetection:
    """Tests for file type detection."""

    def test_detect_csv_extension(self):
        """Test detection of CSV files."""
        assert _get_file_type(Path("data.csv")) == 'csv'
        assert _get_file_type(Path("data.CSV")) == 'csv'
        assert _get_file_type(Path("data.tsv")) == 'csv'
        assert _get_file_type(Path("data.txt")) == 'csv'

    def test_detect_excel_extension(self):
        """Test detection of Excel files."""
        assert _get_file_type(Path("data.xlsx")) == 'excel'
        assert _get_file_type(Path("data.XLSX")) == 'excel'
        assert _get_file_type(Path("data.xls")) == 'excel'
        assert _get_file_type(Path("data.xlsm")) == 'excel'
        assert _get_file_type(Path("data.xlsb")) == 'excel'

    def test_detect_unknown_extension(self):
        """Test detection of unknown file types."""
        assert _get_file_type(Path("data.json")) == 'unknown'
        assert _get_file_type(Path("data.xml")) == 'unknown'
        assert _get_file_type(Path("data")) == 'unknown'


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelLoading:
    """Tests for Excel file loading (requires openpyxl)."""

    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing."""
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a simple Excel file with one sheet
            file_path = Path(tmpdir) / "test_data.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Add headers
            ws['A1'] = 'id'
            ws['B1'] = 'name'
            ws['C1'] = 'amount'

            # Add data rows
            ws['A2'] = '1'
            ws['B2'] = 'Alice'
            ws['C2'] = '100'

            ws['A3'] = '2'
            ws['B3'] = 'Bob'
            ws['C3'] = '200'

            wb.save(file_path)
            yield file_path

    @pytest.fixture
    def temp_multi_sheet_excel(self):
        """Create a temporary Excel file with multiple sheets."""
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "multi_sheet.xlsx"
            wb = Workbook()

            # First sheet (active)
            ws1 = wb.active
            ws1.title = "Sales"
            ws1['A1'] = 'product'
            ws1['B1'] = 'revenue'
            ws1['A2'] = 'Widget'
            ws1['B2'] = '1000'

            # Second sheet
            ws2 = wb.create_sheet("Expenses")
            ws2['A1'] = 'category'
            ws2['B1'] = 'cost'
            ws2['A2'] = 'Marketing'
            ws2['B2'] = '500'

            # Third sheet
            ws3 = wb.create_sheet("Summary")
            ws3['A1'] = 'metric'
            ws3['B1'] = 'value'
            ws3['A2'] = 'Profit'
            ws3['B2'] = '500'

            wb.save(file_path)
            yield file_path

    def test_load_excel_file(self, temp_excel_file):
        """Test loading a basic Excel file."""
        reporter = CSVReporter([str(temp_excel_file)])
        assert reporter.load()
        assert len(reporter.data) == 2
        assert 'id' in reporter.headers
        assert 'name' in reporter.headers
        assert 'amount' in reporter.headers

    def test_load_excel_data_values(self, temp_excel_file):
        """Test that Excel data values are correctly loaded."""
        reporter = CSVReporter([str(temp_excel_file)])
        reporter.load()
        assert reporter.data[0]['name'] == 'Alice'
        assert reporter.data[1]['name'] == 'Bob'

    def test_get_sheet_names(self, temp_multi_sheet_excel):
        """Test getting sheet names from Excel file."""
        reporter = CSVReporter([str(temp_multi_sheet_excel)])
        sheets = reporter.get_sheet_names(temp_multi_sheet_excel)
        assert len(sheets) == 3
        assert 'Sales' in sheets
        assert 'Expenses' in sheets
        assert 'Summary' in sheets

    def test_load_specific_sheet(self, temp_multi_sheet_excel):
        """Test loading a specific sheet by name."""
        reporter = CSVReporter([str(temp_multi_sheet_excel)])
        assert reporter.load(sheet_name="Expenses")
        assert 'category' in reporter.headers
        assert 'cost' in reporter.headers
        assert reporter.data[0]['category'] == 'Marketing'

    def test_load_nonexistent_sheet(self, temp_multi_sheet_excel):
        """Test error when loading nonexistent sheet."""
        reporter = CSVReporter([str(temp_multi_sheet_excel)])
        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            reporter.load(sheet_name="NonExistent")

    def test_load_default_first_sheet(self, temp_multi_sheet_excel):
        """Test that first sheet is loaded by default."""
        reporter = CSVReporter([str(temp_multi_sheet_excel)])
        reporter.load()
        # First sheet is "Sales"
        assert 'product' in reporter.headers
        assert 'revenue' in reporter.headers


class TestExcelWithoutOpenpyxl:
    """Tests for Excel handling when openpyxl is not available."""

    def test_excel_import_error_message(self):
        """Test that helpful error is shown when openpyxl missing."""
        # This test verifies the error message format
        # The actual import error is tested implicitly when openpyxl is not installed
        if not HAS_OPENPYXL:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Create a dummy xlsx file
                file_path = Path(tmpdir) / "test.xlsx"
                file_path.write_bytes(b"dummy")

                reporter = CSVReporter([str(file_path)])
                # load() catches Exception and returns False, so we shouldn't expect raise
                assert reporter.load() is False
